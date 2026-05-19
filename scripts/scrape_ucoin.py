import subprocess, re, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

year = int(sys.argv[1]) if len(sys.argv) > 1 else 2026
_raw_country = sys.argv[2].lower() if len(sys.argv) > 2 else None

CURRENT_YEAR = 2026  # año en curso; fallback del end_year del slug. Revisar cada enero.

# Alias de nombres en español → slug ucoin.net
COUNTRY_ALIASES = {
    "alemania":    "germany",
    "españa":      "spain",
    "austria":     "austria",
    "francia":     "france",
    "italia":      "italy",
    "portugal":    "portugal",
    "belgica":     "belgium",
    "bélgica":     "belgium",
    "holanda":     "netherlands",
    "paises bajos":"netherlands",
    "finlandia":   "finland",
    "irlanda":     "ireland",
    "grecia":      "greece",
    "luxemburgo":  "luxembourg",
    "eslovenia":   "slovenia",
    "chipre":      "cyprus",
    "malta":       "malta",
    "eslovaquia":  "slovakia",
    "estonia":     "estonia",
    "letonia":     "latvia",
    "lituania":    "lithuania",
    "croacia":     "croatia",
    "bulgaria":    "bulgaria",
    "andorra":     "andorra",
    "monaco":      "monaco",
    "mónaco":      "monaco",
    "san marino":  "san_marino",
    "vaticano":    "vatican_city",
    "ciudad del vaticano": "vatican_city",
}

# Las 25 entidades que emiten euros: 21 estados UE de la eurozona
# (Bulgaria entró el 2026-01-01) + 4 microestados con acuerdo monetario.
ALL_EUROZONE = [
    "germany", "france", "italy", "spain", "netherlands", "belgium",
    "austria", "portugal", "finland", "ireland", "greece", "luxembourg",
    "slovenia", "cyprus", "malta", "slovakia", "estonia", "latvia",
    "lithuania", "croatia", "bulgaria",
    "andorra", "monaco", "san_marino", "vatican_city",
]

# Sin país, o "all", se procesa toda la eurozona.
_mapped = COUNTRY_ALIASES.get(_raw_country, _raw_country) if _raw_country else None
run_all = _mapped in (None, "all")
target_country = None if run_all else _mapped

FACE_VALUES   = ["1 Céntimo", "2 Céntimos", "5 Céntimos", "10 Céntimos",
                 "20 Céntimos", "50 Céntimos", "1 Euro", "2 Euros"]
DENOM_LABELS  = ["1ct", "2ct", "5ct", "10ct", "20ct", "50ct", "1€", "2€"]

# ── Estilos Excel ─────────────────────────────────────────────────────
BG_HEADER  = PatternFill("solid", fgColor="1B2A4A")
BG_COUNTRY = PatternFill("solid", fgColor="E8EFF8")
BG_NUMBER  = PatternFill("solid", fgColor="D4EDDA")
BG_PLUS    = PatternFill("solid", fgColor="FFF3CD")
BG_PENDING = PatternFill("solid", fgColor="E9ECEF")
BG_MINUS   = PatternFill("solid", fgColor="FFFFFF")

FT_HEADER  = Font(bold=True, color="F5E8C7")
FT_COUNTRY = Font(bold=True)
FT_TIPO    = Font(bold=True)
FT_NUMBER  = Font(bold=True, color="155724")
FT_PLUS    = Font(bold=True, color="856404")
FT_PENDING = Font(italic=True, color="495057")
FT_MINUS   = Font(color="AAAAAA")

CENTER = Alignment(horizontal="center", vertical="center")

# ── Detección de bloqueo de IP ────────────────────────────────────────
def abort_if_ip_blocked(html):
    """ucoin bloquea la IP tras demasiadas peticiones y sirve una página
    'Oops! ... your IP is blocked' con una cuenta atrás. Si la detectamos,
    abortamos de inmediato: seguir pidiendo solo alarga el bloqueo (escala) y
    produce datos basura. Mejor parar y dejar enfriar la IP unas horas."""
    low = (html or "").lower()
    if "is blocked" in low and "your ip" in low:
        countdown = re.search(r"\d{1,2}:\d{2}:\d{2}", html)
        wait = f"  Cuenta atrás que muestra ucoin: {countdown.group(0)}." if countdown else ""
        print("\n🚫 ucoin ha BLOQUEADO esta IP por exceso de peticiones." + wait)
        print("   Abortado. Espera unas horas (mejor al día siguiente) y vuelve a")
        print("   ejecutar la skill UNA sola vez, sin lanzar peticiones extra: la")
        print("   skill ya se autorregula con sus esperas. No depurar a ráfagas.")
        sys.exit(2)

# ── Safari fetch ──────────────────────────────────────────────────────
first_page = [True]

def fetch_safari(url, url_keyword, extra_delay=3):
    initial_delay = 10 if first_page[0] else 6
    first_page[0] = False
    script = f'''tell application "Safari"
  activate
  open location "{url}"
  delay {initial_delay}
  set waitCount to 0
  repeat
    delay 1
    set waitCount to waitCount + 1
    try
      set currentURL to URL of front document
      if currentURL contains "{url_keyword}" then exit repeat
    end try
    if waitCount >= 15 then exit repeat
  end repeat
  delay {extra_delay}
  return source of front document
end tell'''
    result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, timeout=60)
    abort_if_ip_blocked(result.stdout)
    return result.stdout

# ── URL builder ────────────────────────────────────────────────────────
def build_url_candidates(slug_prefix, start_year, end_year, tid):
    # El end_year guardado va primero (es el correcto). Como fallback se
    # prueban CURRENT_YEAR y vecinos por si ucoin actualizó el slug.
    end_years = [end_year]
    for ey in sorted({CURRENT_YEAR, year, year - 1, year + 1}, reverse=True):
        if ey not in end_years:
            end_years.append(ey)
    return [
        f"https://es.ucoin.net/coin/{slug_prefix}-{start_year}-{ey}/?tid={tid}"
        for ey in end_years
    ]

# ── Parse mintage table ────────────────────────────────────────────────
def parse_mintage(html, target_year):
    """Returns dict with unc_raw/bu_raw/proof_raw (raw strings from ucoin),
    or None if target_year not found in the table."""
    mintage_match = re.search(
        r'id=["\']mintage["\'][^>]*>(.*?)</table>', html, re.DOTALL | re.IGNORECASE
    )
    if not mintage_match:
        return None
    table = mintage_match.group(1)
    has_mint_col = bool(re.search(r'<th[^>]*>\s*Marca\s*</th>', table, re.IGNORECASE))

    rows_raw = re.findall(
        r'<tr[^>]*><td[^>]*><strong>(\d{4})</strong></td>(.*?)</tr>', table, re.DOTALL
    )
    for year_str, cells_html in rows_raw:
        if int(year_str) != target_year:
            continue
        cells = [
            re.sub(r'<[^>]+>', '', c).strip().replace('\xa0', '')
            for c in re.findall(r'<td[^>]*>(.*?)</td>', cells_html, re.DOTALL)
        ]
        if has_mint_col and len(cells) >= 3:
            mint_mark, unc, bu = cells[0], cells[1], cells[2]
            proof = cells[3] if len(cells) > 3 else None
        elif len(cells) >= 2:
            mint_mark, unc, bu = None, cells[0], cells[1]
            proof = cells[2] if len(cells) > 2 else None
        else:
            continue
        return {
            'mint_mark': mint_mark,
            'unc_raw':   unc,
            'bu_raw':    bu,
            'proof_raw': proof,
        }
    return None  # year not found in table

# ── Format raw value for display ──────────────────────────────────────
def format_display(raw):
    """None → None (pending), '-'/'' → '-', '+' → '+', number → '1.234.567'"""
    if raw is None:
        return None
    s = raw.strip()
    if not s or s == '-':
        return '-'
    if '+' in s:
        return '+'
    digits = re.sub(r'[^0-9]', '', s)
    if digits:
        return f"{int(digits):,}".replace(',', '.')
    return s

# ── Auto-descubrimiento de tids ───────────────────────────────────────
DENOM_SLUGS = [
    ("1-euro-cent",  "1 Céntimo"),
    ("2-euro-cent",  "2 Céntimos"),
    ("5-euro-cent",  "5 Céntimos"),
    ("10-euro-cent", "10 Céntimos"),
    ("20-euro-cent", "20 Céntimos"),
    ("50-euro-cent", "50 Céntimos"),
    ("1-euro",       "1 Euro"),
    ("2-euro",       "2 Euros"),
]

def _fetch_catalog(url):
    """Carga una página de catálogo de ucoin en Safari y devuelve el HTML (SSR)."""
    script = f'''tell application "Safari"
  activate
  open location "{url}"
  delay 11
  return source of front document
end tell'''
    result = subprocess.run(['osascript', '-e', script], capture_output=True, text=True, timeout=40)
    abort_if_ip_blocked(result.stdout)
    return result.stdout

def discover_tids(country):
    """Descubre los tids de las 8 denominaciones de un país.
    Paso 1: catálogo general → ID del período 'Unión Europea (Euro)'.
    Paso 2: ese período → URLs /coin/ con tid de cada denominación."""
    print(f"  → Descubriendo tids para {country}...")

    catalog_html = _fetch_catalog(f"https://es.ucoin.net/catalog/?country={country}")
    period_m = re.search(
        rf'country={country}&amp;period=(\d+)"\s+title="[^"]*Euro[^"]*"',
        catalog_html
    )
    if not period_m:
        print(f"  ✗ No se encontró el período Euro para {country}")
        return []
    euro_period = period_m.group(1)

    period_html = _fetch_catalog(
        f"https://es.ucoin.net/catalog/?country={country}&period={euro_period}"
    )
    urls = re.findall(rf'href="/coin/({country}-[^"]*?)/\?tid=(\d+)"', period_html)
    if not urls:
        print(f"  ✗ No se encontraron tids para {country} (período {euro_period})")
        return []

    denoms = []
    for denom_slug, face_value in DENOM_SLUGS:
        candidates = []
        for slug_path, tid in urls:
            # slug_path ej: france-1-euro-2022-2026 — fullmatch evita que '1-euro'
            # capture por error '1-euro-cent'
            if re.fullmatch(rf'{country}-{denom_slug}-\d{{4}}-\d{{4}}', slug_path):
                m = re.search(r'-(\d{4})-(\d{4})$', slug_path)
                candidates.append((int(m.group(1)), int(m.group(2)), int(tid)))
        if candidates:
            # Diseño más reciente = mayor año de fin
            candidates.sort(key=lambda x: (x[1], x[0]), reverse=True)
            start_year, end_year, tid = candidates[0]
            denoms.append((f"{country}-{denom_slug}", start_year, end_year, tid, face_value))
            print(f"    ✓ {face_value}: tid={tid} ({start_year}-{end_year})")
        else:
            print(f"    ✗ {face_value}: no encontrado")

    return denoms

# ── Main scraping loop ────────────────────────────────────────────────
results = []  # list of {pais, denominacion, unc_raw, bu_raw, proof_raw, status}
# status: 'ok' = datos encontrados, 'pending' = año no publicado aún / URL no encontrada

countries_to_run = ALL_EUROZONE if run_all else [target_country]

for country in countries_to_run:
    denoms = discover_tids(country)
    if not denoms:
        print(f"⚠ No se pudo configurar: {country}")
        continue
    print(f"\n▶ {country.upper()} — year {year}")

    for slug_prefix, start_year, end_year, tid, face_value in denoms:
        candidates = build_url_candidates(slug_prefix, start_year, end_year, tid)
        # Reintento: la 1ª candidata (la URL correcta) se reprueba al final
        # por si el primer intento falló por timing de carga de Safari
        candidates = candidates + [candidates[0]]
        found = False

        for url in candidates:
            html = fetch_safari(url, slug_prefix)
            title_m = re.search(r'<title>([^<]+)</title>', html)
            title = title_m.group(1) if title_m else ''

            if 'Page Not Found' in title or 'Not Found' in title or len(html) < 20000:
                continue

            mintage_data = parse_mintage(html, year)

            if mintage_data is None and 'mintage' not in html.lower():
                continue

            if mintage_data:
                unc = format_display(mintage_data['unc_raw'])
                bu  = format_display(mintage_data['bu_raw'])
                prf = format_display(mintage_data['proof_raw'])
                print(f"  ✓ {face_value}: unc={unc}, bu={bu}, proof={prf}")
                results.append({
                    'pais': country, 'denominacion': face_value,
                    'unc_raw': mintage_data['unc_raw'],
                    'bu_raw':  mintage_data['bu_raw'],
                    'proof_raw': mintage_data['proof_raw'],
                    'status': 'ok',
                })
            else:
                print(f"  ⚠ {face_value}: página OK pero sin fila para {year} → pendiente de actualización")
                results.append({
                    'pais': country, 'denominacion': face_value,
                    'unc_raw': None, 'bu_raw': None, 'proof_raw': None,
                    'status': 'pending',
                })
            found = True
            break

        if not found:
            print(f"  ✗ {face_value}: no working URL found → pendiente de actualización")
            results.append({
                'pais': country, 'denominacion': face_value,
                'unc_raw': None, 'bu_raw': None, 'proof_raw': None,
                'status': 'pending',
            })

# ── Construir Excel ───────────────────────────────────────────────────
# Agrupar por país (los dict de Python 3.7+ mantienen orden de inserción)
country_data = {}
for r in results:
    c = r['pais']
    if c not in country_data:
        country_data[c] = {}
    country_data[c][r['denominacion']] = r

wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"Tiradas {year}"

# Fila 1: título
ws.merge_cells('A1:J1')
c = ws['A1']
c.value = f"Tiradas de monedas Euro {year}  —  Unc: circulación  |  BU: sets coleccionista  |  Proof  |  +: acuñada, tirada pendiente"
c.font = Font(bold=True, size=12, color="F5E8C7")
c.fill = BG_HEADER
c.alignment = CENTER
ws.row_dimensions[1].height = 28

# Fila 2: cabeceras columnas
headers = ['País', 'Tipo'] + DENOM_LABELS
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = h
    cell.font = FT_HEADER
    cell.fill = BG_HEADER
    cell.alignment = CENTER
ws.row_dimensions[2].height = 22

TIPOS = [('Unc', 'unc_raw'), ('BU', 'bu_raw'), ('Proof', 'proof_raw')]

current_row = 3
for country, denom_dict in country_data.items():
    start_row = current_row

    for tipo_label, raw_key in TIPOS:
        # Columna B: tipo
        tc = ws.cell(row=current_row, column=2)
        tc.value = tipo_label
        tc.font = FT_TIPO
        tc.alignment = CENTER
        tc.fill = PatternFill("solid", fgColor="F0F4FA")

        # Columnas C-J: denominaciones
        for col_idx, face_value in enumerate(FACE_VALUES, 3):
            row_data = denom_dict.get(face_value)
            cell = ws.cell(row=current_row, column=col_idx)
            cell.alignment = CENTER

            if row_data is None or row_data['status'] == 'pending':
                cell.value = 'pendiente de actualización'
                cell.font = FT_PENDING
                cell.fill = BG_PENDING
            else:
                raw = row_data.get(raw_key)
                display = format_display(raw)
                if display is None:
                    cell.value = 'pendiente de actualización'
                    cell.font = FT_PENDING
                    cell.fill = BG_PENDING
                elif display == '-':
                    cell.value = '—'
                    cell.font = FT_MINUS
                    cell.fill = BG_MINUS
                elif display == '+':
                    cell.value = '+'
                    cell.font = FT_PLUS
                    cell.fill = BG_PLUS
                else:
                    cell.value = display
                    cell.font = FT_NUMBER
                    cell.fill = BG_NUMBER

        ws.row_dimensions[current_row].height = 18
        current_row += 1

    # Columna A: país (fusionada)
    ws.merge_cells(f'A{start_row}:A{current_row - 1}')
    pc = ws.cell(row=start_row, column=1)
    pc.value = country.capitalize()
    pc.font = FT_COUNTRY
    pc.fill = BG_COUNTRY
    pc.alignment = CENTER

# Anchos de columna
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 7
for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws.column_dimensions[col].width = 16

# ── Guardar ────────────────────────────────────────────────────────────
suffix = f"_{target_country}" if target_country else ""
output_path = f"/Users/dariodelapoza/Downloads/ucoin_{year}{suffix}.xlsx"
wb.save(output_path)

print(f"\n✅ {len(results)} filas → {output_path}")
